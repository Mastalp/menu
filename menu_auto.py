# menu omerlo 2.0 using openpyxl instead of csv
# par Laurent-Philippe Roy-Lemaire - lp.roylemaire@gmail.com
#
from googletrans import Translator
from openpyxl import load_workbook
from datetime import datetime
import time
import sys
import os

class colors:
    HEADER = '\033[95m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'

menu_placard = '''
+---------------------------------+
| M E N U   O M E R L O   A U T O |
+---------------------------------+
MENU_AUTOMATIQUE POUR OMERLO V 2.0 par LPRL
'''

os.system('color') # intializing terminal for colored text

translator = Translator() # translator object from googletrans==3.1.0a0

# FUNCTIONS

# returns translated menu_item with googletrans .translate method
def menu_translate(menu_item):
    global progress
    progress += 1
    return (translator.translate(menu_item, dest='en')).text
    
# prints a progress bar (prints over itself)
def progress_bar(progress, total):
    percent = 100 * (progress / float(total))
    bar = '█' * int(percent) + '-' * (100 - int(percent))
    print(f"{colors.HEADER}\r|{bar}| {percent:.2f}%{colors.ENDC}", end="\r")

# returns doc name from date object
def save_as_name(date_object):
    months = ['Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin', 
    'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']
    name = f"Menu {str(date_object.day)} \
        {str(months[int(date_object.month) - 1])} {str(date_object.year)}.xlsx"
    return name

# error handler
def error_handler():
    time.sleep(1)
    print(f"{colors.WARNING}RÉFÉREZ VOUS AU MANUEL AU BESOIN. \
    CETTE FENÊTRE SE FERMERA AUTOMATIQUEMENT {colors.ENDC}")
    time.sleep(7)
    sys.exit(1) # *** EXIT PROGRAM ***

# looping thru SRC and DEST tuples, GETTING and SETTING values using menu list
def menu_auto(src_cells, dest_cells):
    menu = [] # will contain 140 items to be placed in template

    # READ, looping thru src cell tuples, getting and translating cell values
    try:
        for row in src_cells:
            for i in range(len(row)):
                if row[i].value != None and len(row[i].value) > 2: # != et, ou
                    menu_item_f = (row[i].value).strip().capitalize()
                    menu.append(menu_item_f)
                    menu_item_f_t = menu_translate(menu_item_f)
                    menu.append(menu_item_f_t)
                    progress_bar(progress, total)
    except:
        print(f"{colors.FAIL}ERREUR ! Il y a un problème dans \
{src_excel_doc} ! REVOIR LE FORMATAGE!{colors.ENDC}" )
        error_handler()


    # WRITE, looping thru dest cell tuples, setting values
    try:
        for row in dest_cells:
            for i in range(len(row)):
                row[i].value = menu[0] # item at index 0
                menu.pop(0) # popping the stack of delicious items
    except:
        print(f"{colors.FAIL}ERREUR ! Il y a des cases vides dans \
{src_excel_doc} !{colors.ENDC}" )
        error_handler()


### MAIN ###

# get input for src name ? *** TO DO ***
src_excel_doc = 'entrée.xlsx'

# src for save as
dest_excel_doc = 'menu_template.xlsx'

# loading src excel doc with openpyxl w/ error handling
try:
    src_doc = load_workbook(src_excel_doc)
    src_feuille = src_doc.active
except:
    print(f"{colors.FAIL}ERREUR ! Le fichier \
{src_excel_doc} n'est pas dans le répertoire!{colors.ENDC}")
    error_handler()

# loading dest excel doc with openpyxl w/ error handling
try:
    dest_doc = load_workbook(dest_excel_doc)
    dest_feuille = dest_doc.active
except:
    print(f"{colors.FAIL}ERREUR ! Le fichier \
{dest_excel_doc} n'est pas dans le répertoire!{colors.ENDC}")
    error_handler()

# DATE using datetime w/ error handling
# THE DATE NEEDS TO BE A STRING IN SRC EXCEL DOC FOR DATETIME.STRPTIME()
# https://www.programiz.com/python-programming/datetime/strptime
# see Format Code List
try:
    date_cell = dest_feuille['A1']
    date_object = datetime.strptime(src_feuille['A2'].value, "%m/%d/%Y")
    date_cell.value = date_object
except:
    print(f"{colors.FAIL}ERREUR ! \
La date est incorrecte dans {src_excel_doc}!{colors.ENDC}")
    error_handler()

# NAME using save_as_name func and date_object
document_name = save_as_name(date_object)

# SRC and DEST cells from openpyxl ranges as tuples
src_cells = src_feuille['B4':'H20']
dest_cells = dest_feuille['F4':'S13']

# vars for progress bar
progress, total = 0, 70 # items to be translated (10/j, 7j)

# user communication
print(f"{colors.HEADER}{menu_placard}{colors.ENDC}")
time.sleep(2)
print(f"{colors.HEADER}LE PROGRAMME VA COMMENCER...{colors.ENDC}")
time.sleep(1)

# calling main func with openpyxl cell ranges
menu_auto(src_cells, dest_cells)

# SAVING DOC AS using save_as_name function
dest_doc.save(document_name)

print('\r') # print following UNDER progress bar
print(f"{colors.OKGREEN}TRAVAIL TERMINÉ{colors.ENDC}")
time.sleep(1)
print(f"{colors.OKGREEN}LE MENU '{colors.FAIL}{document_name}{colors.ENDC}' \
{colors.OKGREEN}EST MAINTENANT DANS LE RÉPERTOIRE ACTIF{colors.ENDC}")
time.sleep(2)
print(f"{colors.WARNING}CETTE FENÊTRE SE FERMERA AUTOMATIQUEMENT{colors.ENDC}")
time.sleep(7)