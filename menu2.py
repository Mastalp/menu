# menu 2.0 using openpyxl instead of csv
from googletrans import Translator
from openpyxl import load_workbook
from datetime import datetime
import time

translator = Translator()

# FUNCTIONS

# returns translated menu_item with googletrans
def menu_translate(menu_item):
    global progress
    progress += 1
    return (translator.translate(menu_item, dest='en')).text
    
# prints a progress bar
def progress_bar(progress, total):
    percent = 100 * (progress / float(total))
    bar = '█' * int(percent) + '-' * (100 - int(percent))
    print(f"\r|{bar}| {percent:.2f}%", end="\r")

# returns doc name from date object
def doc_name(date_object):
    months = ['Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin', 
    'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']
    name = 'Menu ' + str(date_object.day) + ' ' + \
    str(months[int(date_object.month) - 1]) + ' ' + \
        str(date_object.year) + '.xlsx'
    return name

# looping thru tuples, GETTING and SETTING values
def menu_auto(src_cells, dest_cells):
    menu = [] # will contain 140 items to be placed in template

    # READ, looping thru src cell tuples, getting and translating cell values
    for row in src_cells:
        for i in range(len(row)):
            if row[i].value != None and len(row[i].value) > 2: # != et, ou
                menu_item_f = (row[i].value).strip().capitalize()
                menu.append(menu_item_f)
                menu_item_f_t = menu_translate(menu_item_f)
                menu.append(menu_item_f_t)
                progress_bar(progress, total)

    # WRITE, looping thru dest cell tuples, setting values
    for row in dest_cells:
        for i in range(len(row)):
            row[i].value = menu[0] # item at index 0
            menu.pop(0) # popping the stack of delicious items


# MAIN

# get input for src name ? *** TO DO ***
src_excel_doc = 'entrée.xlsx'

# src for save as
dest_excel_doc = 'menu_template.xlsx'

# loading excel docs with openpyxl
src_doc = load_workbook(src_excel_doc)
src_feuille = src_doc.active
dest_doc = load_workbook(dest_excel_doc)
dest_feuille = dest_doc.active

# DATE using datetime
date_cell = dest_feuille['A1']
date_object = datetime.strptime(src_feuille['A2'].value, "%d-%m-%Y")
date_cell.value = date_object

# NAME using doc_name func
document_name = doc_name(date_object)

# SRC and DEST cell tuples from openpyxl ranges
src_cells = src_feuille['B4':'H20']
dest_cells = dest_feuille['F4':'S13']

# vars for progress bar
progress, total = 0, 70 # items to be translated (10/j, 7j)

print("MENU_AUTOMATIQUE POUR OMERLO V 2.0 par LP Roy-LEMAIRE")
time.sleep(2)
print("LE PROGRAMME VA COMMENCER...")
time.sleep(1)

# CALLING MAIN FUNC ***
menu_auto(src_cells, dest_cells)

# SAVING DOC AS
dest_doc.save(document_name)

print('\r')

print("TRAVAIL TERMINÉ")
time.sleep(1)
print("LE MENU EST MAINTENANT DANS LE RÉPERTOIRE ACTIF")
time.sleep(2)
print("CETTE FENÊTRE SE FERMERA AUTOMATIQUEMENT")
time.sleep(5)