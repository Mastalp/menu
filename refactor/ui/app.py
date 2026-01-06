import threading
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json
from openpyxl import load_workbook
from core import MenuFiller, GoogleTranslator, TkProgressBar
from core.extractor import ExtractorFactory
from core.interfaces import *

class MenuAutoApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Menu Auto")
        self.geometry("800x300") 

        # ─── grid config ──────────────────────────────────────────
        # 3 cols: only the middle one stretches
        self.columnconfigure(0, weight=0)
        self.columnconfigure(1, weight=1)
        self.columnconfigure(2, weight=0)

        # rows: header/separator/banner fixed, 
        # inputs fixed, progress+filler stretch, button fixed
        self.rowconfigure(0, weight=0)  # in-app title
        self.rowconfigure(1, weight=0)  # line
        self.rowconfigure(2, weight=0)  # banner
        self.rowconfigure(3, weight=0)  # source picker
        self.rowconfigure(4, weight=0)  # template picker
        self.rowconfigure(5, weight=0)  # mode selector
        self.rowconfigure(6, weight=1)  # progress + empty space
        self.rowconfigure(7, weight=0)  # generate button

        # ─── in-app title + separator ─────────────────────────────
        tk.Label(self, text="Menu Automatique pour Omerlo, v3.1.1 par LPRL", font=("Helvetica", 16, "bold")) \
          .grid(row=0, column=0, columnspan=3, pady=(10,0))
        ttk.Separator(self, orient="horizontal") \
          .grid(row=1, column=0, columnspan=3, sticky="ew", pady=5)

        # ─── banner ────────────────────────────────────────────────
        tk.Label(self, text="Choisissez le document excel source et le template", font=("Helvetica", 14)) \
          .grid(row=2, column=0, columnspan=3, pady=(0,20))

        # ─── source picker ─────────────────────────────────────────
        tk.Label(self, text="Menu source:") \
          .grid(row=3, column=0, sticky="e", padx=5, pady=5)
        self.src_path = tk.StringVar()
        
        tk.Entry(self, textvariable=self.src_path) \
          .grid(row=3, column=1, sticky="ew", padx=5, pady=5)
        tk.Button(self, text="choisir", command=self.browse_src) \
          .grid(row=3, column=2, padx=5, pady=5)

        # ─── template picker ───────────────────────────────────────
        tk.Label(self, text="Menu template:") \
          .grid(row=4, column=0, sticky="e", padx=5, pady=5)
        self.tpl_path = tk.StringVar()
        tk.Entry(self, textvariable=self.tpl_path) \
          .grid(row=4, column=1, sticky="ew", padx=5, pady=5)
        tk.Button(self, text="choisir", command=self.browse_tpl) \
          .grid(row=4, column=2, padx=5, pady=5)

        # ─── mode selector ─────────────────────────────────────────
        tk.Label(self, text="Mode:") \
          .grid(row=5, column=0, sticky="e", padx=5, pady=5)
        self.template_var = tk.StringVar(value="officiel")
        with open("templates.json") as f:
            choices = list(json.load(f).keys())
        tk.OptionMenu(self, self.template_var, *choices) \
          .grid(row=5, column=1, sticky="w", padx=5, pady=5)

        # ─── progress bar + spacer ─────────────────────────────────
        self.progress = ttk.Progressbar(self, orient="horizontal", mode="determinate")
        self.progress.grid(row=6, column=0, columnspan=3, sticky="ew", padx=5)

        # ─── centered Generate button ──────────────────────────────
        self.generate_btn = tk.Button(
            self,
            text="GÉNÉRER",
            width=20,
            command=self.on_generate
        )
        # no sticky ⇒ stays at its natural size, centered in the span
        self.generate_btn.grid(row=7, column=0, columnspan=3, pady=20)


        # TEST ------------------------------------------- * * * 
        s = self.src_path = "/home/mastalp/Documents/menu/refactor/source_file.xlsx"
        t = self.tpl_path = "/home/mastalp/Documents/menu/refactor/template_omerlo.xlsx"
        e_factory = ExtractorFactory("templates.json")
        extractor = e_factory.get(self.template_var.get())

        src_wb = load_workbook(s)
        dest_wb = load_workbook(t)

        mf = MenuFiller(src_wb, dest_wb, extractor, GoogleTranslator(), TkProgressBar(self.progress))

        mf.run()
        dest_wb.save("/home/mastalp/Documents/menu/refactor/output.xlsx")


        # --------------------


    def browse_src(self):
        path = filedialog.askopenfilename(filetypes=[("Excel","*.xlsx")])
        if path:
            self.src_path.set(path)

    def browse_tpl(self):
        path = filedialog.askopenfilename(filetypes=[("Excel","*.xlsx")])
        if path:
            self.tpl_path.set(path)

    def on_generate(self):
        # Disable button to prevent re-entry
        self.generate_btn.config(state='disabled')

        # Prepare template and determine total items
        
        #factory = TemplateFactory("templates.json")
        #template = factory.get(self.template_var.get())
        #src_wb = load_workbook(self.src_path.get())
        #total = len(template.read_items(src_wb))
        #self.progress['value'] = 0
        #self.progress['maximum'] = total

        

        # Start translation/write in background thread
        #worker = threading.Thread(
        #    target=self._worker, 
        #    args=(template,), 
        #    daemon=True
        #)
        
        #worker.start()

    def _worker(self, template: IMenuTemplate):
        try:
            src_wb = load_workbook(self.src_path.get())
            tpl_wb = load_workbook(self.tpl_path.get())

            filler = MenuFiller(
                template,
                GoogleTranslator(),
                TkProgressBar(self.progress)
            )

            # Execute the fill process and get the file name 
            out_name = filler.run(src_wb, tpl_wb)

            # Save result 
            tpl_wb.save(out_name)

            self.after(0, lambda: messagebox.showinfo("✅ Success", f"Saved: {out_name}"))
        except Exception as e:
            self.after(0, lambda err=e: messagebox.showerror("❌ Error", str(err)))
        finally:
            # Re-enable the generate button
            self.after(0, lambda: self.generate_btn.config(state='normal'))

    

if __name__ == "__main__":
    app = MenuAutoApp()
    app.mainloop()
